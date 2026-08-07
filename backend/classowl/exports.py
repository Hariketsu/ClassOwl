from __future__ import annotations

import re
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, ConfigDict
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .storage import DocumentStore, NotFoundError


pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))


class ExportOptions(BaseModel):
    model_config = ConfigDict(extra="ignore", strict=True)

    classes: list[str] = []
    fileName: str = ""
    title: str = ""
    showTeacher: bool = True
    showNotes: bool = False
    showBiweekly: bool = True
    sheetLayout: Literal[
        "每班一个工作表", "每年级一个工作表", "班级总表"
    ] = "每班一个工作表"
    includeStats: bool = False
    paper: Literal["A4", "A3"] = "A4"
    orientation: Literal["横向", "纵向"] = "横向"
    pagination: Literal[
        "每班一页", "每年级一页", "连续排版"
    ] = "每班一页"


class ExportRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    format: Literal["excel", "pdf"]
    options: ExportOptions = ExportOptions()
    targetPath: str


@dataclass
class ExportJob:
    id: str
    status: Literal["queued", "running", "done", "error"] = "queued"
    progress: int = 0
    message: str = "等待导出"
    path: str | None = None

    def response(self) -> dict[str, Any]:
        response: dict[str, Any] = {
            "status": self.status,
            "progress": self.progress,
            "message": self.message,
        }
        if self.path is not None:
            response["path"] = self.path
        return response


class ExportService:
    def __init__(self, store: DocumentStore):
        self.store = store
        self.jobs: dict[str, ExportJob] = {}
        self._lock = threading.Lock()

    def start(
        self,
        plan_id: str,
        export_format: Literal["excel", "pdf"],
        options: ExportOptions,
        target_path: str,
    ) -> ExportJob:
        document = self.store.get_document(plan_id)["doc"]
        job = ExportJob(id=str(uuid4()))
        with self._lock:
            self.jobs[job.id] = job
        threading.Thread(
            target=self._export,
            args=(job, document, export_format, options, target_path),
            daemon=True,
        ).start()
        return job

    def get(self, job_id: str) -> ExportJob:
        try:
            return self.jobs[job_id]
        except KeyError as error:
            raise NotFoundError("导出任务不存在") from error

    def _export(
        self,
        job: ExportJob,
        document: dict[str, Any],
        export_format: Literal["excel", "pdf"],
        options: ExportOptions,
        target_path: str,
    ) -> None:
        with self._lock:
            job.status = "running"
            job.progress = 10
            job.message = f"正在生成 {export_format.upper()}"
        try:
            if export_format == "pdf":
                _write_pdf(document, options, Path(target_path))
            else:
                _write_excel(document, options, Path(target_path))
            with self._lock:
                job.status = "done"
                job.progress = 100
                job.message = "导出完成"
                job.path = target_path
        except Exception as error:
            with self._lock:
                job.status = "error"
                job.progress = 100
                job.message = str(error)


def _class_label(school_class: dict[str, Any]) -> str:
    return f"{school_class['grade']}{school_class['name']}"


def _sheet_name(name: str, workbook: Workbook) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "_", name).strip("'")[:31] or "课表"
    candidate = base
    number = 2
    while candidate in workbook.sheetnames:
        suffix = f" ({number})"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        number += 1
    return candidate


def _biweekly_tag(document: dict[str, Any], course_id: str) -> str:
    for item in document["biweekly"]:
        if course_id not in (item["courseA"], item["courseB"]):
            continue
        return "单" if (
            item["oddCourseId"] == course_id or item["courseA"] == course_id
        ) else "双"
    return ""


def _selected_classes(
    document: dict[str, Any], options: ExportOptions
) -> list[dict[str, Any]]:
    selected = set(options.classes)
    return [
        school_class
        for school_class in document["classes"]
        if not selected or school_class["id"] in selected
    ]


def _timetable_data(
    document: dict[str, Any],
    school_class: dict[str, Any],
    options: ExportOptions,
    title: str,
) -> list[list[str]]:
    rows = [[title, "", *(day["label"] for day in document["days"])]]
    course_names = {
        course["id"]: course["name"] for course in document["courses"]
    }
    teacher_names = {
        teacher["id"]: teacher["name"] for teacher in document["teachers"]
    }
    placements = {
        (item["dayId"], item["periodId"]): item
        for item in document["placements"]
        if item["classId"] == school_class["id"]
    }
    for period in document["periods"]:
        row = [period["band"], period["label"] or f"{period['id']}节"]
        for day in document["days"]:
            placement = placements.get((day["id"], period["id"]))
            if placement is None:
                row.append("")
                continue
            course = course_names.get(
                placement["courseId"], placement["courseId"]
            )
            if options.showBiweekly:
                tag = _biweekly_tag(document, placement["courseId"])
                if tag:
                    course = f"{course}（{tag}）"
            teacher = teacher_names.get(placement["teacherId"], "")
            row.append(
                f"{course}\n{teacher}"
                if options.showTeacher and teacher
                else course
            )
        rows.append(row)
    return rows


def _write_timetable(
    sheet: Worksheet,
    document: dict[str, Any],
    school_class: dict[str, Any],
    options: ExportOptions,
    start_column: int = 1,
    title: str | None = None,
) -> None:
    periods = document["periods"]
    data = _timetable_data(
        document,
        school_class,
        options,
        title if title is not None else options.title,
    )
    for row_offset, row in enumerate(data):
        for column_offset, value in enumerate(row):
            sheet.cell(
                row_offset + 1,
                start_column + column_offset,
                value,
            )

    header_row = 1
    sheet.merge_cells(
        start_row=header_row,
        start_column=start_column,
        end_row=header_row,
        end_column=start_column + 1,
    )
    corner = sheet.cell(header_row, start_column)
    corner.font = Font(bold=True)
    corner.alignment = Alignment(horizontal="center", vertical="center")
    for offset in range(2, len(data[0])):
        cell = sheet.cell(header_row, start_column + offset)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(
        min_row=2,
        max_row=len(data),
        min_col=start_column + 2,
        max_col=start_column + len(data[0]) - 1,
    ):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    run_start = 0
    for index in range(1, len(periods) + 1):
        if (
            index < len(periods)
            and periods[index]["band"] == periods[run_start]["band"]
        ):
            continue
        if index > run_start + 1:
            sheet.merge_cells(
                start_row=run_start + 2,
                start_column=start_column,
                end_row=index + 1,
                end_column=start_column,
            )
        run_start = index


def _write_stats(
    workbook: Workbook,
    document: dict[str, Any],
    class_ids: set[str],
) -> None:
    sheet = workbook.create_sheet(_sheet_name("统计", workbook))
    sheet.append(["教师", "总课时数"])
    counts: dict[str, int] = {}
    for placement in document["placements"]:
        if placement["classId"] in class_ids:
            teacher_id = placement["teacherId"]
            counts[teacher_id] = counts.get(teacher_id, 0) + 1
    for teacher in document["teachers"]:
        sheet.append([teacher["name"], counts.get(teacher["id"], 0)])


def _write_excel(
    document: dict[str, Any], options: ExportOptions, path: Path
) -> None:
    classes = _selected_classes(document, options)
    workbook = Workbook()
    workbook.remove(workbook.active)
    table_width = len(document["days"]) + 2

    if options.sheetLayout == "每班一个工作表":
        for school_class in classes:
            sheet = workbook.create_sheet(
                _sheet_name(_class_label(school_class), workbook)
            )
            _write_timetable(sheet, document, school_class, options)
    elif options.sheetLayout == "每年级一个工作表":
        grades: list[str] = []
        for school_class in classes:
            if school_class["grade"] not in grades:
                grades.append(school_class["grade"])
        for grade in grades:
            sheet = workbook.create_sheet(_sheet_name(grade, workbook))
            grade_classes = [
                item for item in classes if item["grade"] == grade
            ]
            for index, school_class in enumerate(grade_classes):
                _write_timetable(
                    sheet,
                    document,
                    school_class,
                    options,
                    1 + index * (table_width + 1),
                    f"{options.title}（{_class_label(school_class)}）",
                )
    else:
        sheet = workbook.create_sheet(_sheet_name("班级总表", workbook))
        for index, school_class in enumerate(classes):
            _write_timetable(
                sheet,
                document,
                school_class,
                options,
                1 + index * (table_width + 1),
                f"{options.title}（{_class_label(school_class)}）",
            )

    if options.includeStats:
        _write_stats(
            workbook, document, {item["id"] for item in classes}
        )
    workbook.save(path)


def _pdf_table(
    document: dict[str, Any],
    school_class: dict[str, Any],
    options: ExportOptions,
    width: float,
) -> Table:
    class_name = _class_label(school_class)
    title = f"{options.title}（{class_name}）" if options.title else class_name
    data = _timetable_data(document, school_class, options, title)
    day_width = width * 0.8 / len(document["days"])
    table = Table(
        data,
        colWidths=[width * 0.08, width * 0.12]
        + [day_width] * len(document["days"]),
        repeatRows=1,
    )
    style = [
        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
        ("SPAN", (0, 0), (1, 0)),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF7")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#8A94A6")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEADING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    periods = document["periods"]
    run_start = 0
    for index in range(1, len(periods) + 1):
        if (
            index < len(periods)
            and periods[index]["band"] == periods[run_start]["band"]
        ):
            continue
        if index > run_start + 1:
            style.append(("SPAN", (0, run_start + 1), (0, index)))
        run_start = index
    table.setStyle(TableStyle(style))
    return table


def _write_pdf(
    document: dict[str, Any], options: ExportOptions, path: Path
) -> None:
    pagesize = A3 if options.paper == "A3" else A4
    if options.orientation == "横向":
        pagesize = landscape(pagesize)
    pdf = SimpleDocTemplate(
        str(path),
        pagesize=pagesize,
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    classes = _selected_classes(document, options)
    tables = [
        _pdf_table(document, school_class, options, pdf.width)
        for school_class in classes
    ]
    story: list[Any] = []
    if options.pagination == "每年级一页":
        grades = list(dict.fromkeys(item["grade"] for item in classes))
        for grade_index, grade in enumerate(grades):
            grade_tables: list[Any] = []
            for table in (
                tables[index]
                for index, item in enumerate(classes)
                if item["grade"] == grade
            ):
                if grade_tables:
                    grade_tables.append(Spacer(1, 12))
                grade_tables.append(table)
            story.append(KeepTogether(grade_tables))
            if grade_index < len(grades) - 1:
                story.append(PageBreak())
    else:
        for index, table in enumerate(tables):
            story.append(table)
            if index < len(tables) - 1:
                story.append(
                    PageBreak()
                    if options.pagination == "每班一页"
                    else Spacer(1, 12)
                )
    pdf.build(story)
