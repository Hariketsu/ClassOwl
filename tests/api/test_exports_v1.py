from __future__ import annotations

import re
import time
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from classowl.app import create_app


TOKEN = "test-token"


@pytest.fixture
def client(tmp_path: Path):
    with TestClient(create_app(TOKEN, tmp_path)) as test_client:
        test_client.headers["X-ClassOwl-Token"] = TOKEN
        yield test_client


@pytest.fixture
def plan_id(client: TestClient) -> str:
    document = {
        "schemeName": "导出测试",
        "days": [
            {"id": 1, "label": "星期一", "short": "一"},
            {"id": 2, "label": "星期二", "short": "二"},
        ],
        "periods": [
            {
                "id": 1,
                "label": "第一节",
                "band": "上午",
                "start": "08:00",
                "end": "08:40",
            },
            {
                "id": 2,
                "label": "",
                "band": "下午",
                "start": "14:00",
                "end": "14:40",
            },
        ],
        "teachers": [
            {"id": "t1", "name": "王老师"},
            {"id": "t2", "name": "李老师"},
        ],
        "classes": [
            {"id": "c1", "grade": "一年级", "name": "1班", "room": "101"},
            {"id": "c2", "grade": "一年级", "name": "2班", "room": "102"},
            {"id": "c3", "grade": "二年级", "name": "1班", "room": "201"},
        ],
        "courses": [
            {"id": "k1", "name": "语文", "biweekly": False},
            {"id": "k2", "name": "数学", "biweekly": False},
        ],
        "gradeCourses": {"一年级": ["k1", "k2"], "二年级": ["k1"]},
        "matrix": {
            "c1": {"k1": {"hours": 1, "teacherId": "t1"}},
            "c2": {"k2": {"hours": 1, "teacherId": "t2"}},
            "c3": {"k1": {"hours": 1, "teacherId": "t1"}},
        },
        "linked": [],
        "biweekly": [],
        "combined": [],
        "layered": [],
        "venues": [],
        "rules": [],
        "placements": [
            {
                "id": "p1",
                "classId": "c1",
                "courseId": "k1",
                "teacherId": "t1",
                "dayId": 1,
                "periodId": 1,
                "source": "manual",
                "locked": False,
            },
            {
                "id": "p2",
                "classId": "c2",
                "courseId": "k2",
                "teacherId": "t2",
                "dayId": 2,
                "periodId": 2,
                "source": "manual",
                "locked": False,
            },
        ],
        "park": [],
        "scheduleStatus": "ready",
    }
    plan = client.post("/api/v1/plans").json()
    saved = client.put(
        f"/api/v1/plans/{plan['id']}/doc",
        json={"baseRev": 0, "doc": document, "checkpoint": None},
    )
    assert saved.status_code == 200, saved.text
    return plan["id"]


def export(
    client: TestClient,
    plan_id: str,
    path: Path,
    format: str = "excel",
    **options: object,
) -> dict:
    request_options = {
        "classes": ["c1", "c2"],
        "fileName": "课表",
        "title": "2026 秋季课表",
        "showTeacher": True,
        "showNotes": False,
        "showBiweekly": True,
        "sheetLayout": "每班一个工作表",
        "includeStats": False,
        **options,
    }
    started = client.post(
        f"/api/v1/plans/{plan_id}/exports",
        json={
            "format": format,
            "options": request_options,
            "targetPath": str(path),
        },
    )
    assert started.status_code == 202, started.text
    assert started.json()["jobId"]
    for _ in range(200):
        job = client.get(
            f"/api/v1/exports/{started.json()['jobId']}"
        ).json()
        if job["status"] not in {"queued", "running"}:
            return job
        time.sleep(0.01)
    pytest.fail("导出任务未在时限内结束")


def pdf_size(path: Path) -> tuple[float, float]:
    match = re.search(
        rb"/MediaBox\s*\[\s*0\s+0\s+([\d.]+)\s+([\d.]+)\s*\]",
        path.read_bytes(),
    )
    assert match
    return float(match.group(1)), float(match.group(2))


def pdf_page_count(path: Path) -> int:
    return len(re.findall(rb"/Type\s*/Page(?!s)", path.read_bytes()))


def test_export_job_writes_file(client: TestClient, plan_id: str, tmp_path: Path):
    target = tmp_path / "schedule.xlsx"

    job = export(client, plan_id, target)

    assert job["status"] == "done"
    assert job["path"] == str(target)
    assert target.exists()


def test_one_sheet_per_class(client: TestClient, plan_id: str, tmp_path: Path):
    target = tmp_path / "classes.xlsx"
    export(client, plan_id, target)

    workbook = load_workbook(target)

    assert workbook.sheetnames == ["一年级1班", "一年级2班"]


def test_cell_contains_course_and_teacher(
    client: TestClient, plan_id: str, tmp_path: Path
):
    target = tmp_path / "content.xlsx"
    export(client, plan_id, target)

    sheet = load_workbook(target)["一年级1班"]

    assert sheet["C2"].value == "语文\n王老师"


def test_show_teacher_false(client: TestClient, plan_id: str, tmp_path: Path):
    target = tmp_path / "no-teacher.xlsx"
    export(client, plan_id, target, showTeacher=False)

    value = load_workbook(target)["一年级1班"]["C2"].value

    assert value == "语文"
    assert "王老师" not in value


def test_master_layout_has_one_timetable_sheet(
    client: TestClient, plan_id: str, tmp_path: Path
):
    target = tmp_path / "master.xlsx"
    export(client, plan_id, target, sheetLayout="班级总表")

    assert load_workbook(target).sheetnames == ["班级总表"]


def test_stats_contains_teacher_hours(
    client: TestClient, plan_id: str, tmp_path: Path
):
    target = tmp_path / "stats.xlsx"
    export(client, plan_id, target, includeStats=True)

    stats = load_workbook(target)["统计"]
    values = list(stats.values)

    assert ("王老师", 1) in values
    assert ("李老师", 1) in values


def test_timetable_header_matches_preview(
    client: TestClient, plan_id: str, tmp_path: Path
):
    target = tmp_path / "header.xlsx"
    export(client, plan_id, target)

    sheet = load_workbook(target)["一年级1班"]

    assert [sheet.cell(1, column).value for column in range(1, 5)] == [
        "2026 秋季课表",
        None,
        "星期一",
        "星期二",
    ]
    assert [sheet.cell(2, column).value for column in range(1, 3)] == [
        "上午",
        "第一节",
    ]
    assert [sheet.cell(3, column).value for column in range(1, 3)] == [
        "下午",
        "2节",
    ]


def test_missing_plan_returns_404(client: TestClient, tmp_path: Path):
    response = client.post(
        "/api/v1/plans/missing/exports",
        json={
            "format": "excel",
            "options": {},
            "targetPath": str(tmp_path / "missing.xlsx"),
        },
    )

    assert response.status_code == 404


def test_unknown_option_is_ignored(
    client: TestClient, plan_id: str, tmp_path: Path
):
    target = tmp_path / "unknown-option.xlsx"

    job = export(client, plan_id, target, paper="A4")

    assert job["status"] == "done"
    assert target.exists()


def test_export_does_not_change_document(
    client: TestClient, plan_id: str, tmp_path: Path
):
    before = client.get(f"/api/v1/plans/{plan_id}/doc").json()

    job = export(client, plan_id, tmp_path / "read-only.xlsx")
    after = client.get(f"/api/v1/plans/{plan_id}/doc").json()

    assert job["status"] == "done"
    assert after == before


def test_one_sheet_per_grade_places_classes_side_by_side(
    client: TestClient, plan_id: str, tmp_path: Path
):
    target = tmp_path / "grades.xlsx"
    export(
        client,
        plan_id,
        target,
        classes=["c1", "c2", "c3"],
        sheetLayout="每年级一个工作表",
    )

    workbook = load_workbook(target)

    assert workbook.sheetnames == ["一年级", "二年级"]
    assert workbook["一年级"]["A1"].value == "2026 秋季课表（一年级1班）"
    assert workbook["一年级"]["F1"].value == "2026 秋季课表（一年级2班）"


def test_pdf_export_job_writes_pdf(
    client: TestClient, plan_id: str, tmp_path: Path
):
    target = tmp_path / "schedule.pdf"

    job = export(client, plan_id, target, format="pdf")

    assert job["status"] == "done"
    assert job["path"] == str(target)
    assert target.read_bytes().startswith(b"%PDF")


def test_pdf_a4_landscape_is_wider_than_tall(
    client: TestClient, plan_id: str, tmp_path: Path
):
    target = tmp_path / "a4-landscape.pdf"
    export(
        client,
        plan_id,
        target,
        format="pdf",
        paper="A4",
        orientation="横向",
    )

    width, height = pdf_size(target)

    assert width > height


def test_pdf_a3_is_larger_than_a4(
    client: TestClient, plan_id: str, tmp_path: Path
):
    a4 = tmp_path / "a4.pdf"
    a3 = tmp_path / "a3.pdf"
    export(client, plan_id, a4, format="pdf", paper="A4")
    export(client, plan_id, a3, format="pdf", paper="A3")

    a4_width, a4_height = pdf_size(a4)
    a3_width, a3_height = pdf_size(a3)

    assert a3_width > a4_width
    assert a3_height > a4_height


def test_pdf_portrait_is_taller_than_wide(
    client: TestClient, plan_id: str, tmp_path: Path
):
    target = tmp_path / "portrait.pdf"
    export(client, plan_id, target, format="pdf", orientation="纵向")

    width, height = pdf_size(target)

    assert height > width


def test_pdf_pagination_controls_page_breaks(
    client: TestClient, plan_id: str, tmp_path: Path
):
    per_class = tmp_path / "per-class.pdf"
    continuous = tmp_path / "continuous.pdf"
    common = {"format": "pdf", "classes": ["c1", "c2", "c3"]}
    export(
        client,
        plan_id,
        per_class,
        pagination="每班一页",
        **common,
    )
    export(
        client,
        plan_id,
        continuous,
        pagination="连续排版",
        **common,
    )

    per_class_pages = pdf_page_count(per_class)
    continuous_pages = pdf_page_count(continuous)

    assert per_class_pages >= 3
    assert continuous_pages < per_class_pages


def test_pdf_references_builtin_cid_font(
    client: TestClient, plan_id: str, tmp_path: Path
):
    target = tmp_path / "cid-font.pdf"
    export(client, plan_id, target, format="pdf")

    assert b"STSong" in target.read_bytes()


def test_pdf_ignores_excel_option(
    client: TestClient, plan_id: str, tmp_path: Path
):
    target = tmp_path / "unknown-option.pdf"

    job = export(
        client,
        plan_id,
        target,
        format="pdf",
        sheetLayout="班级总表",
    )

    assert job["status"] == "done"
    assert target.exists()


def test_pdf_missing_plan_returns_404(client: TestClient, tmp_path: Path):
    response = client.post(
        "/api/v1/plans/missing/exports",
        json={
            "format": "pdf",
            "options": {},
            "targetPath": str(tmp_path / "missing.pdf"),
        },
    )

    assert response.status_code == 404
